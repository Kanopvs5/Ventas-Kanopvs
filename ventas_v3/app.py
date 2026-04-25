from flask import Flask, render_template, request, jsonify, send_file, session, redirect
import os, io
from datetime import datetime
from supabase import create_client

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "kiddoshoes2026secretkey")

SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://bbzmwneywbbgamnttply.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "sb_publishable_kDvKdK8S-AZxiGRnBRqpHw_Tjky7Uhe")
ADMIN_EMAIL  = os.environ.get("ADMIN_EMAIL", "")

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Auth ──────────────────────────────────────────────────────────────────────

def get_user():
    token = session.get("access_token")
    if not token:
        return None
    try:
        r = sb.auth.get_user(token)
        return r.user
    except:
        return None

def is_admin(user):
    if not user or not ADMIN_EMAIL:
        return False
    return user.email == ADMIN_EMAIL

# ── Rutas ─────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    user = get_user()
    if not user:
        return redirect("/login")
    return render_template("index.html", user_email=user.email, is_admin=is_admin(user))

@app.route("/login")
def login_page():
    return render_template("login.html")

@app.route("/api/auth/login", methods=["POST"])
def api_login():
    data = request.get_json()
    try:
        r = sb.auth.sign_in_with_password({"email": data["email"].strip(), "password": data["password"]})
        session["access_token"]  = r.session.access_token
        session["refresh_token"] = r.session.refresh_token
        session["user_email"]    = r.user.email
        session["user_id"]       = str(r.user.id)
        return jsonify({"ok": True, "email": r.user.email})
    except:
        return jsonify({"error": "Correo o contraseña incorrectos"}), 401

@app.route("/api/auth/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})

# ── API Productos ─────────────────────────────────────────────────────────────

@app.route("/api/productos", methods=["GET"])
def api_get_productos():
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    r = sb.table("productos").select("*").execute()
    return jsonify(r.data)

@app.route("/api/productos", methods=["POST"])
def api_crear_producto():
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    data = request.get_json()
    producto = {
        "id":        data["nombre"].strip(),
        "nombre":    data["nombre"].strip(),
        "categoria": data.get("categoria","").strip(),
        "codigo":    data.get("codigo","").strip(),
        "marca":     data.get("marca","").strip(),
        "precio":    float(data.get("precio", 0)),
        "ganancia":  float(data.get("ganancia", 0)),
        "color":     data.get("color", "#ffffff"),
    }
    r = sb.table("productos").insert(producto).execute()
    return jsonify(r.data[0] if r.data else producto), 201

@app.route("/api/productos/<path:pid>", methods=["PUT"])
def api_editar_producto(pid):
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    data = request.get_json()
    update = {
        "categoria": data.get("categoria","").strip(),
        "codigo":    data.get("codigo","").strip(),
        "marca":     data.get("marca","").strip(),
        "precio":    float(data.get("precio", 0)),
        "ganancia":  float(data.get("ganancia", 0)),
        "color":     data.get("color", "#ffffff"),
    }
    r = sb.table("productos").update(update).eq("id", pid).execute()
    return jsonify(r.data[0] if r.data else update)

@app.route("/api/productos/<path:pid>", methods=["DELETE"])
def api_eliminar_producto(pid):
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    sb.table("productos").delete().eq("id", pid).execute()
    return jsonify({"ok": True})

# ── API Subproductos ──────────────────────────────────────────────────────────

@app.route("/api/subproductos", methods=["GET"])
def api_get_subproductos():
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    pid = request.args.get("producto_id")
    q = sb.table("subproductos").select("*")
    if pid:
        q = q.eq("producto_id", pid)
    r = q.execute()
    return jsonify(r.data)

@app.route("/api/subproductos", methods=["POST"])
def api_crear_subproducto():
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    data = request.get_json()
    item = {
        "producto_id":  data["producto_id"],
        "modelo":       data.get("modelo","").strip(),
        "descripcion":  data.get("descripcion","").strip(),
        "talla_min":    int(data.get("talla_min", 18)),
        "talla_max":    int(data.get("talla_max", 42)),
        "tallas":       data.get("tallas", {}),
    }
    r = sb.table("subproductos").insert(item).execute()
    return jsonify(r.data[0] if r.data else item), 201

@app.route("/api/subproductos/<int:sid>", methods=["PUT"])
def api_editar_subproducto(sid):
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    data = request.get_json()
    update = {
        "modelo":      data.get("modelo","").strip(),
        "descripcion": data.get("descripcion","").strip(),
        "talla_min":   int(data.get("talla_min", 18)),
        "talla_max":   int(data.get("talla_max", 42)),
        "tallas":      data.get("tallas", {}),
    }
    r = sb.table("subproductos").update(update).eq("id", sid).execute()
    return jsonify(r.data[0] if r.data else update)

@app.route("/api/subproductos/<int:sid>", methods=["DELETE"])
def api_eliminar_subproducto(sid):
    user = get_user()
    if not user or not is_admin(user): return jsonify({"error": "Sin permisos"}), 403
    sb.table("subproductos").delete().eq("id", sid).execute()
    return jsonify({"ok": True})

# ── API Ventas ────────────────────────────────────────────────────────────────

@app.route("/api/ventas", methods=["GET"])
def api_get_ventas():
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    if is_admin(user):
        r = sb.table("ventas").select("*").order("id", desc=True).execute()
    else:
        r = sb.table("ventas").select("*").eq("user_email", user.email).order("id", desc=True).execute()
    return jsonify(r.data)

@app.route("/api/ventas", methods=["POST"])
def api_registrar_venta():
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    data     = request.get_json()
    pid      = data["producto_id"]
    cant     = int(data["cantidad"])
    fecha    = data.get("fecha") or datetime.now().strftime("%Y-%m-%d")
    talla    = data.get("talla")
    subprod_id = data.get("subproducto_id")
    modelo   = data.get("modelo","")

    # Validar fecha para empleados
    hoy = datetime.now().strftime("%Y-%m-%d")
    if not is_admin(user) and fecha != hoy:
        return jsonify({"error": "Solo el admin puede registrar ventas en otras fechas"}), 403

    pr = sb.table("productos").select("*").eq("id", pid).execute()
    if not pr.data:
        return jsonify({"error": "Producto no encontrado"}), 404
    producto = pr.data[0]

    # Descontar inventario si tiene subproducto y talla
    if subprod_id and talla:
        sr = sb.table("subproductos").select("*").eq("id", subprod_id).execute()
        if sr.data:
            sub = sr.data[0]
            tallas = sub.get("tallas", {})
            talla_key = str(talla)
            stock_actual = tallas.get(talla_key, 0)
            if stock_actual < cant:
                return jsonify({"error": f"Stock insuficiente para talla {talla}. Disponible: {stock_actual}"}), 400
            tallas[talla_key] = stock_actual - cant
            sb.table("subproductos").update({"tallas": tallas}).eq("id", subprod_id).execute()

    venta = {
        "producto_id":     pid,
        "nombre":          producto["nombre"],
        "categoria":       producto["categoria"],
        "codigo":          producto["codigo"],
        "marca":           producto["marca"],
        "precio_unitario": producto["precio"],
        "ganancia":        producto["ganancia"],
        "cantidad":        cant,
        "total":           producto["precio"] * cant,
        "ganancia_total":  producto["ganancia"] * cant,
        "fecha":           fecha,
        "hora":            datetime.now().strftime("%H:%M"),
        "user_id":         session.get("user_id"),
        "user_email":      user.email,
        "talla":           talla,
        "subproducto_id":  subprod_id,
        "modelo":          modelo,
    }
    r = sb.table("ventas").insert(venta).execute()
    return jsonify(r.data[0] if r.data else venta), 201

@app.route("/api/ventas/<int:vid>", methods=["DELETE"])
def api_eliminar_venta(vid):
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    sb.table("ventas").delete().eq("id", vid).execute()
    return jsonify({"ok": True})

# ── API Reportes ──────────────────────────────────────────────────────────────

@app.route("/api/reportes", methods=["GET"])
def api_reportes():
    user = get_user()
    if not user: return jsonify({"error": "No autenticado"}), 401
    mes  = request.args.get("mes", datetime.now().strftime("%Y-%m"))
    
    if is_admin(user):
        r = sb.table("ventas").select("*").like("fecha", f"{mes}%").execute()
    else:
        r = sb.table("ventas").select("*").eq("user_email", user.email).like("fecha", f"{mes}%").execute()
    
    ventas = r.data
    total  = sum(v["total"] for v in ventas)
    ganancia = sum(v.get("ganancia_total", 0) for v in ventas)
    
    # Top productos
    conteo = {}
    for v in ventas:
        n = v["nombre"]
        conteo[n] = conteo.get(n, {"cantidad": 0, "total": 0})
        conteo[n]["cantidad"] += v["cantidad"]
        conteo[n]["total"]    += v["total"]
    top = sorted(conteo.items(), key=lambda x: x[1]["cantidad"], reverse=True)[:10]

    # Ventas por día
    por_dia = {}
    for v in ventas:
        d = v["fecha"]
        por_dia[d] = por_dia.get(d, 0) + v["total"]
    por_dia_list = [{"fecha": k, "total": v} for k, v in sorted(por_dia.items())]

    return jsonify({
        "total": total,
        "ganancia": ganancia,
        "cantidad_ventas": len(ventas),
        "top_productos": [{"nombre": k, **v} for k, v in top],
        "por_dia": por_dia_list,
    })

# ── Exportar Excel ────────────────────────────────────────────────────────────

@app.route("/api/exportar/excel")
def exportar_excel():
    user = get_user()
    if not user: return redirect("/login")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"error":"pip install openpyxl"}), 500

    if is_admin(user):
        ventas = sb.table("ventas").select("*").order("id", desc=False).execute().data
    else:
        ventas = sb.table("ventas").select("*").eq("user_email", user.email).order("id", desc=False).execute().data

    productos = sb.table("productos").select("*").execute().data
    wb = Workbook()

    hf   = PatternFill("solid", start_color="1a1a18")
    hfnt = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    nfnt = Font(name="Arial", size=10)
    alt  = PatternFill("solid", start_color="F5F4F0")
    bs   = Side(style="thin", color="E2E0D8")
    brd  = Border(left=bs, right=bs, top=bs, bottom=bs)
    ctr  = Alignment(horizontal="center", vertical="center")

    def set_header(ws, headers, widths):
        ws.merge_cells(f"A1:{chr(64+len(headers))}1")
        ws["A1"] = f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True, name="Arial", size=12)
        ws["A1"].alignment = ctr
        ws.row_dimensions[1].height = 28
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = hfnt; c.fill = hf; c.alignment = ctr; c.border = brd
            ws.column_dimensions[c.column_letter].width = w
        ws.row_dimensions[2].height = 20

    ws = wb.active
    ws.title = "VENTAS"
    set_header(ws, ["Fecha","Producto","Modelo","Talla","Marca","Precio","Cant.","Total","Ganancia","Vendedor"],
                   [14,20,12,8,18,14,8,14,14,22])
    total_v = gan_v = 0
    for i, v in enumerate(ventas):
        r = i + 3
        fill = alt if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
        for col, val in enumerate([v["fecha"],v["nombre"],v.get("modelo",""),v.get("talla",""),
                v.get("marca",""),v["precio_unitario"],v["cantidad"],v["total"],
                v.get("ganancia_total",0),v.get("user_email","")], 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = nfnt; c.fill = fill; c.border = brd; c.alignment = ctr
        total_v += v["total"]; gan_v += v.get("ganancia_total",0)
        ws.row_dimensions[r].height = 16
    tr = len(ventas) + 3
    for col, val in enumerate(["","TOTAL","","","","","",total_v,gan_v,""], 1):
        c = ws.cell(row=tr, column=col, value=val)
        c.font = Font(bold=True, name="Arial", size=10)
        c.fill = PatternFill("solid", start_color="EAF3DE"); c.border = brd; c.alignment = ctr

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fname)

@app.route("/api/exportar/pdf")
def exportar_pdf():
    user = get_user()
    if not user: return redirect("/login")
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    except ImportError:
        return jsonify({"error":"pip install reportlab"}), 500

    if is_admin(user):
        ventas = sb.table("ventas").select("*").order("id", desc=False).execute().data
    else:
        ventas = sb.table("ventas").select("*").eq("user_email", user.email).order("id", desc=False).execute().data

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    t_style = ParagraphStyle("t", parent=styles["Heading1"], fontSize=14, spaceAfter=4)
    s_style = ParagraphStyle("s", parent=styles["Normal"], fontSize=8, spaceAfter=12)
    elems = [Paragraph("Reporte de Ventas — KiddoShoes", t_style),
             Paragraph(f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}", s_style)]

    hdr  = ["Fecha","Producto","Modelo","Talla","Marca","Precio","Cant.","Total","Ganancia"]
    rows = [hdr]
    total_v = gan_v = 0
    for v in ventas:
        rows.append([v["fecha"],v["nombre"],v.get("modelo",""),str(v.get("talla","")),
                     v.get("marca",""),f"${v['precio_unitario']:,.0f}",str(v["cantidad"]),
                     f"${v['total']:,.0f}",f"${v.get('ganancia_total',0):,.0f}"])
        total_v += v["total"]; gan_v += v.get("ganancia_total",0)
    rows.append(["","TOTAL","","","","","",f"${total_v:,.0f}",f"${gan_v:,.0f}"])

    cw = [2.2*cm,3.8*cm,2*cm,1.5*cm,3*cm,2.5*cm,1.5*cm,2.5*cm,2.5*cm]
    t  = Table(rows, colWidths=cw, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1a1a18")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),7),
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.HexColor("#F5F4F0"),colors.white]),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#E2E0D8")),
        ("ROWHEIGHT",(0,0),(-1,-1),0.55*cm),
        ("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#EAF3DE")),
    ]))
    elems.append(t)
    doc.build(elems)
    buf.seek(0)
    fname = f"reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fname)

if __name__ == "__main__":
    print("🚀 Servidor iniciado en http://127.0.0.1:5000")
    app.run(host="0.0.0.0", port=5000, debug=False)
